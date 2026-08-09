"""Flinders ISCEV Control ERG parser.

Parses ``ISCEV Control ERG Flinders University.xlsx`` into typed
SubjectRecord / VisitRecord / SessionRecord / WaveformRecord objects plus
Waveform arrays.

Source semantics observed:

- ``Flinders Normal`` sheet: one row per (subject, protocol, eye) with
  a/b-wave times and amplitudes, OP amplitude/time, age, sex, eye.
  ``Test`` in {LA3, 30Hz, DA001, DA3, DA10}; healthy controls only
  (target 0).  Contains None cells and near-duplicate rows; both are
  counted, never silently dropped.
- ``FIGURES`` sheet: stacked blocks; each block is a column pair
  (``ms``, ``uV``) preceded by 13 metadata cells in the same column
  (protocol, subject id, test dates, eye, age).  Traces use 0.512 ms
  steps => 1953.125 Hz.  Blocks whose metadata cells are missing (some
  traces at rows 444+ carry no repeated header) are reported through
  ``iter_figures_blocks`` with ``metadata_missing=True`` and are never
  silently merged into another block.
- ``Summary Stats`` sheet is empty in this export and is ignored.
"""

from __future__ import annotations

import hashlib
import json
from collections.abc import Iterator
from pathlib import Path

import numpy as np
import openpyxl

from .schemas import (
    Dataset,
    Eye,
    Protocol,
    SessionRecord,
    SubjectRecord,
    VisitRecord,
    Waveform,
    WaveformKind,
    WaveformRecord,
)

DATASET = Dataset.FLINDERS

EYE_MAP = {"LeftEye": Eye.LEFT, "RightEye": Eye.RIGHT}

PROTOCOL_MAP = {
    "LA3": Protocol.LA3,
    "30Hz": Protocol.FLICKER_30HZ,
    "DA0.01": Protocol.DA001,
    "DA001": Protocol.DA001,
    "DA3": Protocol.DA3,
    "DA10": Protocol.DA10,
    "OPS": Protocol.OPS,
}

HEADER_ROWS = 13  # rows 0..12 hold block metadata; row 13 is ms/uV
DATA_START = 14


class FlindersParseError(ValueError):
    pass


def _opt_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def make_flinders_subject_id(source_id: str) -> str:
    return f"FLINDERS_{source_id}"


def _visit_id(subject_id: str) -> str:
    return f"{subject_id}_V0"


def _session_id(visit_id: str, protocol: Protocol, eye: Eye) -> str:
    return f"{visit_id}_{protocol.value}_{eye.value}"


def _recording_key(subject_id: str, protocol: str, eye: str) -> str:
    raw = json.dumps(
        {"subject": subject_id, "protocol": protocol, "eye": eye}, sort_keys=True
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:12]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Flinders Normal sheet (tabular features)
# ---------------------------------------------------------------------------


def iter_feature_rows(xlsx_path: str | Path) -> Iterator[dict]:
    """Yield rows of the ``Flinders Normal`` sheet as dicts (deterministic)."""
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb["Flinders Normal"]
        header: list[str] = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c) if c is not None else "" for c in row]
                continue
            if all(c is None for c in row):
                continue
            yield {h: row[j] if j < len(row) else None for j, h in enumerate(header)}
    finally:
        wb.close()


def parse_subject_features(
    row: dict, xlsx_path: Path, source_row: int
) -> tuple[SubjectRecord, VisitRecord, SessionRecord, WaveformRecord | None]:
    """Parse one ``Flinders Normal`` row into typed records.

    The row carries landmark features (a/b times+amps, OP) but no waveform;
    the returned ``WaveformRecord`` is None and the features are kept in the
    subject-level record for join with FIGURES traces.
    """
    raw_id = str(row.get("id") or "").strip()
    if not raw_id:
        raise FlindersParseError(f"missing id at Flinders Normal row {source_row}")
    protocol_raw = str(row.get("Test") or "").strip()
    if protocol_raw not in PROTOCOL_MAP:
        raise FlindersParseError(
            f"unknown protocol {protocol_raw!r} at Flinders Normal row {source_row}"
        )
    protocol = PROTOCOL_MAP[protocol_raw]
    eye_raw = str(row.get("Eye") or "").strip()
    if eye_raw not in EYE_MAP:
        raise FlindersParseError(f"unknown eye {eye_raw!r} at row {source_row}")
    eye = EYE_MAP[eye_raw]

    subject_id = make_flinders_subject_id(raw_id)
    visit_id = _visit_id(subject_id)
    session_id = _session_id(visit_id, protocol, eye)

    age = _opt_float(row.get("age"))
    sex_raw = row.get("sex")
    sex_standardized = str(int(sex_raw)) if sex_raw is not None else None
    subject = SubjectRecord(
        global_subject_id=subject_id,
        dataset=DATASET,
        source_subject_id=raw_id,
        repeat_component_id=None,
        age_years=age,
        sex_raw=str(sex_raw) if sex_raw is not None else None,
        sex_standardized=sex_standardized,
        site="Flinders",
        group_raw="Normal",
        source_checksum=sha256_file(xlsx_path),
    )
    visit = VisitRecord(
        global_visit_id=visit_id,
        global_subject_id=subject_id,
        dataset=DATASET,
        source_record_id=raw_id,
        visit_date=None,
        diagnosis1_raw="Normal",
        diagnosis2_raw=None,
        diagnosis3_raw=None,
        target_binary=0,  # healthy controls are the reference class
        target_multiclass=0,
        target_mapping_version="flinders_labels_v1",
    )
    session = SessionRecord(
        global_session_id=session_id,
        global_visit_id=visit_id,
        dataset=DATASET,
        source_session_index=0,
        session_type=protocol.value,
        acquisition_timestamp_start=None,
        eyes_available=(eye,),
    )

    features = {
        k: row.get(k)
        for k in ("a_time", "a_amp", "b_time", "b_amp", "OP_s_Amp", "OP_s_Time")
    }
    supplied = json.dumps(
        {"source_row": source_row, **features}, sort_keys=True, default=str
    )
    rec = WaveformRecord(
        global_recording_id=(
            f"{subject_id}_{protocol.value}_{eye.value}_feat"
            f"_{_recording_key(raw_id, protocol.value, eye.value)}"
        ),
        global_subject_id=subject_id,
        global_visit_id=visit_id,
        global_session_id=session_id,
        dataset=DATASET,
        protocol=protocol,
        eye=eye,
        stimulus_value=None,
        stimulus_unit="",
        waveform_kind=WaveformKind.ERG,
        source_wave_id=f"{raw_id}_{protocol.value}_{eye.value}",
        source_file=xlsx_path.name,
        source_row_or_column=str(source_row),
        array_key="",
        n_samples=0,
        start_ms=float("nan"),
        end_ms=float("nan"),
        median_dt_ms=float("nan"),
        sampling_rate_hz=float("nan"),
        supplied_features_json=supplied,
    )
    return subject, visit, session, rec


# ---------------------------------------------------------------------------
# FIGURES sheet (waveform traces)
# ---------------------------------------------------------------------------


def iter_figures_blocks(xlsx_path: str | Path) -> Iterator[dict]:
    """Yield one dict per waveform block found in the FIGURES sheet.

    Deterministic scan: every cell equal to ``ms`` whose right neighbor is
    ``uV`` starts a block.  The 13 cells above (same column) are the block
    metadata; when they are absent the block is still yielded with
    ``metadata_missing=True`` so counts stay complete and nothing is
    silently merged.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        ws = wb["FIGURES"]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    height = len(rows)
    max_col = len(rows[0]) if height else 0
    for ms_r in range(height):
        for col in range(max_col - 1):
            if rows[ms_r][col] != "ms" or rows[ms_r][col + 1] != "uV":
                continue
            meta = _block_metadata(rows, ms_r, col)
            time_ms, signal_uv = [], []
            time_ms, signal_uv = [], []
            for i in range(ms_r + 1, height):
                tm, sig = rows[i][col], rows[i][col + 1]
                if isinstance(tm, str) or isinstance(sig, str):
                    break  # next block header begins
                if tm is None and sig is None:
                    break
                if tm is None or sig is None:
                    raise FlindersParseError(
                        f"partial trace at row {i} col {col} (one of ms/uV missing)"
                    )
                time_ms.append(float(tm))
                signal_uv.append(float(sig))
            yield {
                "protocol_raw": meta["protocol"],
                "subject_id": meta["subject_id"],
                "test_date": meta["test_date"],
                "eye_raw": meta["eye"],
                "age": meta["age"],
                "extra_metrics": meta["metrics"],
                "metadata_missing": meta["missing"],
                "time_ms": np.asarray(time_ms, dtype=float),
                "signal_uv": np.asarray(signal_uv, dtype=float),
                "empty": len(time_ms) == 0,
            }


def _block_metadata(rows: list[tuple], ms_row: int, col: int) -> dict:
    """Read the metadata cells above a ``ms``/``uV`` pair in one column."""
    meta: dict = {
        "protocol": None,
        "subject_id": None,
        "test_date": None,
        "eye": None,
        "age": None,
        "metrics": {},
        "missing": False,
    }
    start = ms_row - HEADER_ROWS
    if start < 0:
        meta["missing"] = True
        return meta
    for k, idx in (
        ("protocol", 0),
        ("subject_id", 1),
        ("test_date", 2),
        ("eye", 4),
        ("age", 5),
    ):
        v = rows[start + idx][col]
        meta[k] = str(v).strip() if v is not None else None
    for idx in (6, 7, 8, 9, 10, 11):
        v = rows[start + idx][col]
        if v is not None:
            meta["metrics"][f"metric_{idx}"] = v
    meta["missing"] = meta["protocol"] not in PROTOCOL_MAP
    return meta


def parse_figures_trace(
    block: dict, xlsx_path: Path
) -> tuple[SubjectRecord, VisitRecord, SessionRecord, Waveform | None]:
    """Convert one FIGURES block into typed records plus a Waveform."""
    if block.get("metadata_missing") or block.get("empty"):
        raise FlindersParseError(
            "FIGURES block without metadata or without data; cannot bind to a subject"
        )
    raw_id = block["subject_id"] or ""
    protocol_raw = str(block.get("protocol_raw") or "").strip()
    if protocol_raw not in PROTOCOL_MAP:
        raise FlindersParseError(f"unknown FIGURES protocol {protocol_raw!r}")
    protocol = PROTOCOL_MAP[protocol_raw]
    eye_raw = str(block.get("eye_raw") or "").strip()
    eye = EYE_MAP.get(eye_raw)
    if eye is None:
        raise FlindersParseError(f"unknown FIGURES eye {eye_raw!r}")

    subject_id = make_flinders_subject_id(raw_id)
    visit_id = _visit_id(subject_id)
    session_id = _session_id(visit_id, protocol, eye)

    subject = SubjectRecord(
        global_subject_id=subject_id,
        dataset=DATASET,
        source_subject_id=raw_id,
        repeat_component_id=None,
        age_years=_opt_float(block.get("age")),
        sex_raw=None,
        sex_standardized=None,
        site="Flinders",
        group_raw="Normal",
        source_checksum=sha256_file(xlsx_path),
    )
    visit = VisitRecord(
        global_visit_id=visit_id,
        global_subject_id=subject_id,
        dataset=DATASET,
        source_record_id=raw_id,
        visit_date=block.get("test_date"),
        diagnosis1_raw="Normal",
        diagnosis2_raw=None,
        diagnosis3_raw=None,
        target_binary=0,
        target_multiclass=0,
        target_mapping_version="flinders_labels_v1",
    )
    session = SessionRecord(
        global_session_id=session_id,
        global_visit_id=visit_id,
        dataset=DATASET,
        source_session_index=0,
        session_type=protocol.value,
        acquisition_timestamp_start=block.get("test_date"),
        eyes_available=(eye,),
    )

    time_ms = block["time_ms"]
    signal_uv = block["signal_uv"]
    if time_ms.size != signal_uv.size:
        raise FlindersParseError(
            f"trace length mismatch for {raw_id}/{protocol_raw}: "
            f"{time_ms.size} vs {signal_uv.size}"
        )
    n = time_ms.size
    dt = float(np.median(np.diff(time_ms))) if n > 1 else float("nan")
    rec_id = (
        f"{subject_id}_{protocol.value}_{eye.value}"
        f"_{_recording_key(raw_id, protocol_raw, eye_raw)}"
    )
    record = WaveformRecord(
        global_recording_id=rec_id,
        global_subject_id=subject_id,
        global_visit_id=visit_id,
        global_session_id=session_id,
        dataset=DATASET,
        protocol=protocol,
        eye=eye,
        stimulus_value=None,
        stimulus_unit="",
        waveform_kind=WaveformKind.ERG,
        source_wave_id=f"{raw_id}_{protocol.value}_{eye.value}",
        source_file=xlsx_path.name,
        source_row_or_column="FIGURES",
        array_key=f"{subject_id}/{protocol.value}/{eye.value}",
        n_samples=n,
        start_ms=float(time_ms[0]),
        end_ms=float(time_ms[-1]),
        median_dt_ms=dt,
        sampling_rate_hz=1000.0 / dt if dt and np.isfinite(dt) else float("nan"),
        supplied_features_json=json.dumps(
            {"metrics": block.get("extra_metrics") or {}}, sort_keys=True, default=str
        ),
    )
    return subject, visit, session, Waveform(time_ms=time_ms, signal_uv=signal_uv, record=record)


def iter_flinders(
    xlsx_path: str | Path,
) -> Iterator[
    tuple[SubjectRecord, VisitRecord, SessionRecord, Waveform | None, str]
]:
    """Yield all records from both sheets (Normal rows then FIGURES).

    The fifth element is the source kind: ``features`` or ``trace``.
    """
    path = Path(xlsx_path)
    for i, row in enumerate(iter_feature_rows(path), start=2):
        yield (*parse_subject_features(row, path, i), "features")
    for block in iter_figures_blocks(path):
        if block.get("metadata_missing") or block.get("empty"):
            yield from ()
            continue
        yield (*parse_figures_trace(block, path), "trace")


def summarize_counts(xlsx_path: str | Path) -> dict:
    """Counts used for EXPECTED_COUNTS cross-checking."""
    path = Path(xlsx_path)
    feature_rows = 0
    feature_by_protocol: dict[str, int] = {}
    feature_subjects: set[str] = set()
    traces = empty_traces = missing_meta = 0
    traces_by_protocol: dict[str, int] = {}
    trace_subjects: set[str] = set()
    near_duplicates = 0
    missing_cells = 0
    seen: set[tuple] = set()
    for row in iter_feature_rows(path):
        key = tuple(
            row.get(k) for k in ("id", "Test", "Eye", "a_time", "a_amp", "b_time", "b_amp")
        )
        if key in seen:
            near_duplicates += 1
        seen.add(key)
        missing_cells += sum(
            1 for k in ("a_time", "a_amp", "b_time", "b_amp") if row.get(k) is None
        )
        feature_rows += 1
        proto = str(row.get("Test") or "").strip()
        feature_by_protocol[proto] = feature_by_protocol.get(proto, 0) + 1
        feature_subjects.add(str(row.get("id")))
    for block in iter_figures_blocks(path):
        if block["empty"]:
            empty_traces += 1
            continue
        if block.get("metadata_missing"):
            missing_meta += 1
            continue
        traces += 1
        trace_subjects.add(block["subject_id"])
        p = block["protocol_raw"]
        traces_by_protocol[p] = traces_by_protocol.get(p, 0) + 1
    return {
        "feature_rows": feature_rows,
        "feature_subjects": len(feature_subjects),
        "feature_rows_by_protocol": feature_by_protocol,
        "waveform_traces": traces,
        "empty_traces": empty_traces,
        "metadata_missing_traces": missing_meta,
        "waveform_traces_by_protocol": traces_by_protocol,
        "waveform_subjects": len(trace_subjects),
        "near_duplicate_feature_rows": near_duplicates,
        "missing_feature_cells": missing_cells,
    }
