"""URFU Pediatric and Adults ERG Database parser.

Parses ``01 Appendix 1.xlsx`` from the ``Pediatric and Adults ERG
Database`` folder into typed SubjectRecord / VisitRecord / SessionRecord /
WaveformRecord objects plus Waveform arrays.

Source semantics observed:

- One sheet per protocol: ``Maximum 2.0 ERG Response``, ``Scotopic 2.0
  ERG Response``, ``Photopic 2.0 ERG Response``, ``Photopic 2.0 ERG
  Flicker`` (30 Hz flicker) and ``Oscillatory Potentials``.  Each sheet is
  a single header block: row 2 ``Quantity``, row 3 ``Type``, row 4 ``#``
  (per-column subject ids), an optional ``Diagnosis`` row, an ``Age`` row,
  one or two feature blocks, a ``Graph`` marker and finally a ``Time, ms``
  row with one ``Signal, µV`` column per recording.
- Sampling is 0.5 ms steps => 2000 Hz.  Signal columns end at the last row
  where both time and signal are numeric; trailing rows that only hold time
  values (and padding cells) are ignored.
- The ``Oscillatory Potentials`` sheet carries no Diagnosis row and no
  feature rows; some of its signal columns have empty ``#`` cells (those
  recordings are bound to a derived unlabeled subject id and counted, never
  dropped).
- No eye labels exist anywhere in the database: ``eye`` is always None.
- Feature rows can repeat (a second, partially populated block).  Both
  blocks are kept separately (``features_1`` / ``features_2``) in
  ``supplied_features_json``; non-numeric values (e.g. ``- healthy
  ch...``) count as missing cells.
- The same subject id appears in many columns (both within one sheet and
  across sheets); every column is its own recording and nothing is
  deduplicated.

Explicitly excluded: ``02 Appendix 2.xlsx`` (sheet ``urfu``) is a
scattered fragment dump: sub-traces of a few samples with truncated
strings (``0.0,0.0,0.``) and no time axis or protocol labels; it cannot be
mapped to WaveformRecords and is never loaded.
"""

from __future__ import annotations

import hashlib
import json
from collections.abc import Iterator
from pathlib import Path

import numpy as np
import openpyxl

from .schemas import (
    Dataset,
    Protocol,
    SessionRecord,
    SubjectRecord,
    VisitRecord,
    Waveform,
    WaveformKind,
    WaveformRecord,
)

DATASET = Dataset.URFU

SHEET_ORDER = (
    "Maximum 2.0 ERG Response",
    "Scotopic 2.0 ERG Response",
    "Photopic 2.0 ERG Response",
    "Photopic 2.0 ERG Flicker",
    "Oscillatory Potentials",
)

PROTOCOL_MAP: dict[str, Protocol] = {
    "Maximum 2.0 ERG Response": Protocol.MAXIMUM,
    "Scotopic 2.0 ERG Response": Protocol.SCOTOPIC,
    "Photopic 2.0 ERG Response": Protocol.PHOTOPIC,
    "Photopic 2.0 ERG Flicker": Protocol.FLICKER_30HZ,
    "Oscillatory Potentials": Protocol.OPS,
}

FEATURE_KEYWORDS = (
    "a-wave",
    "b-wave",
    "Amplitude, µ",
    "Latency, ms",
)


class UrfuParseError(ValueError):
    pass


def _cell(rows: list[tuple], r: int, c: int) -> object:
    if r >= len(rows) or c >= len(rows[r]):
        return None
    v = rows[r][c]
    if v == "":
        return None
    return v


def _opt_float(value) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _sheet_short(sheet: str) -> str:
    return {
        "Maximum 2.0 ERG Response": "MAX",
        "Scotopic 2.0 ERG Response": "SCO",
        "Photopic 2.0 ERG Response": "PHO",
        "Photopic 2.0 ERG Flicker": "FLI",
        "Oscillatory Potentials": "OPS",
    }[sheet]


def make_urfu_subject_id(raw_id: str) -> str:
    return f"URFU_{raw_id}"


def _visit_id(subject_id: str) -> str:
    return f"{subject_id}_V0"


def _session_id(visit_id: str, sheet_short: str, col: int) -> str:
    return f"{visit_id}_{sheet_short}_c{col}"


def _recording_key(sheet: str, col: int, subject_id: str) -> str:
    raw = json.dumps(
        {"sheet": sheet, "col": col, "subject": subject_id}, sort_keys=True
    )
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:12]


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Sheet-level parsing
# ---------------------------------------------------------------------------


def iter_urfu_blocks(xlsx_path: str | Path) -> Iterator[dict]:
    """Yield one block dict per (sheet, signal column), deterministic order.

    Keys: ``sheet``, ``column``, ``subject_id`` (may be None), ``diagnosis``,
    ``age``, ``features_1``, ``features_2``, ``time_ms``, ``signal_uv``,
    ``empty``.
    """
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    try:
        sheets = {name: list(wb[name].iter_rows(values_only=True)) for name in wb.sheetnames}
    finally:
        wb.close()
    for sheet in SHEET_ORDER:
        rows = sheets.get(sheet)
        if rows is None:
            continue
        for c in _signal_columns(rows, sheet):
            yield from _extract_block(rows, sheet, c)


def _signal_columns(rows: list[tuple], sheet: str) -> list[int]:
    tr = _find_label_row(rows, 0, len(rows), "Time, ms")
    if tr is None:
        raise UrfuParseError(f"{sheet}: no 'Time, ms' row")
    return [c for c in range(1, len(rows[tr])) if rows[tr][c] == "Signal, µV"]


def _find_label_row(
    rows: list[tuple], r_lo: int, r_hi: int, label: str
) -> int | None:
    for r in range(r_lo, min(r_hi, len(rows))):
        if _cell(rows, r, 0) == label:
            return r
    return None


def _extract_block(rows: list[tuple], sheet: str, col: int) -> tuple[dict]:
    hash_row = _find_label_row(rows, 0, 12, "#")
    if hash_row is None:
        raise UrfuParseError(f"{sheet}: no '#' row")
    raw_id = _cell(rows, hash_row, col)
    subject_id = str(raw_id).strip() if raw_id not in (None, "") else None
    if subject_id is None:
        subject_id = f"UNLABELED_{_sheet_short(sheet)}_c{col}"

    diag_row = _find_label_row(rows, hash_row + 1, hash_row + 4, "Diagnosis")
    age_row = _find_label_row(rows, hash_row + 1, hash_row + 4, "Age")
    if age_row is None:
        raise UrfuParseError(f"{sheet}: no 'Age' row")

    diagnosis_raw: str | None = None
    if diag_row is not None:
        v = _cell(rows, diag_row, col)
        diagnosis_raw = str(v).strip() if v not in (None, "") else None
    age = _opt_float(_cell(rows, age_row, col))

    time_row = _find_label_row(rows, age_row + 1, len(rows), "Time, ms")
    if time_row is None:
        raise UrfuParseError(f"{sheet}: no 'Time, ms' row")

    features_1: dict[str, float | None] = {}
    features_2: dict[str, float | None] = {}
    for r in range(age_row + 1, time_row):
        label = _cell(rows, r, 0)
        if label is None:
            continue
        label_s = str(label).strip()
        if not any(label_s.startswith(k) for k in FEATURE_KEYWORDS):
            continue
        if label_s in features_1:
            features_2[label_s] = _opt_float(_cell(rows, r, col))
        else:
            features_1[label_s] = _opt_float(_cell(rows, r, col))

    time_ms: list[float] = []
    signal_uv: list[float] = []
    for r in range(time_row + 1, len(rows)):
        t = _opt_float(_cell(rows, r, 0))
        s = _opt_float(_cell(rows, r, col))
        if t is None or s is None:
            break
        time_ms.append(t)
        signal_uv.append(s)
    return [
        {
            "sheet": sheet,
            "column": col,
            "subject_id": subject_id,
            "diagnosis": diagnosis_raw,
            "age": age,
            "features_1": features_1,
            "features_2": features_2,
            "time_ms": np.asarray(time_ms, dtype=float),
            "signal_uv": np.asarray(signal_uv, dtype=float),
            "empty": len(time_ms) == 0,
        }
    ]


def parse_urfu_block(
    block: dict, xlsx_path: Path
) -> tuple[SubjectRecord, VisitRecord, SessionRecord, Waveform | None]:
    """Convert one URFU block into typed records plus a Waveform."""
    if block["empty"]:
        raise UrfuParseError("URFU block without data; cannot bind to a subject")
    sheet = block["sheet"]
    col = block["column"]
    raw_id = block["subject_id"] or ""
    protocol = PROTOCOL_MAP[sheet]

    subject_id = make_urfu_subject_id(raw_id)
    visit_id = _visit_id(subject_id)
    session_id = _session_id(visit_id, _sheet_short(sheet), col)

    subject = SubjectRecord(
        global_subject_id=subject_id,
        dataset=DATASET,
        source_subject_id=raw_id,
        repeat_component_id=None,
        age_years=block["age"],
        sex_raw=None,
        sex_standardized=None,
        site="URFU",
        group_raw=sheet,
        source_checksum=sha256_file(xlsx_path),
    )
    visit = VisitRecord(
        global_visit_id=visit_id,
        global_subject_id=subject_id,
        dataset=DATASET,
        source_record_id=f"{sheet}:c{col}",
        visit_date=None,
        diagnosis1_raw=block["diagnosis"],
        diagnosis2_raw=None,
        diagnosis3_raw=None,
        target_binary=None,
        target_multiclass=None,
        target_mapping_version="urfu_v1",
    )
    session = SessionRecord(
        global_session_id=session_id,
        global_visit_id=visit_id,
        dataset=DATASET,
        source_session_index=col,
        session_type=protocol.value,
        acquisition_timestamp_start=None,
        eyes_available=(),
    )

    time_ms = block["time_ms"]
    signal_uv = block["signal_uv"]
    if time_ms.size != signal_uv.size:
        raise UrfuParseError(
            f"trace length mismatch {sheet}:c{col}: {time_ms.size} vs {signal_uv.size}"
        )
    n = time_ms.size
    dt = float(np.median(np.diff(time_ms))) if n > 1 else float("nan")
    record = WaveformRecord(
        global_recording_id=(
            f"{subject_id}_{protocol.value}_c{col}"
            f"_{_recording_key(sheet, col, raw_id)}"
        ),
        global_subject_id=subject_id,
        global_visit_id=visit_id,
        global_session_id=session_id,
        dataset=DATASET,
        protocol=protocol,
        eye=None,
        stimulus_value=None,
        stimulus_unit="",
        waveform_kind=WaveformKind.OP
        if protocol is Protocol.OPS
        else WaveformKind.ERG,
        source_wave_id=f"{sheet}:c{col}",
        source_file=xlsx_path.name,
        source_row_or_column=f"c{col}",
        array_key=f"{subject_id}/{protocol.value}/{col}",
        n_samples=n,
        start_ms=float(time_ms[0]),
        end_ms=float(time_ms[-1]),
        median_dt_ms=dt,
        sampling_rate_hz=1000.0 / dt if dt and np.isfinite(dt) else float("nan"),
        supplied_features_json=json.dumps(
            {
                "sheet": sheet,
                "column": col,
                "features_1": block["features_1"],
                "features_2": block["features_2"],
                "diagnosis": block["diagnosis"],
            },
            sort_keys=True,
            default=str,
        ),
    )
    return subject, visit, session, Waveform(time_ms=time_ms, signal_uv=signal_uv, record=record)


def iter_urfu(
    xlsx_path: str | Path,
) -> Iterator[
    tuple[SubjectRecord, VisitRecord, SessionRecord, Waveform | None]
]:
    path = Path(xlsx_path)
    for block in iter_urfu_blocks(path):
        yield parse_urfu_block(block, path)


def summarize_counts(xlsx_path: str | Path) -> dict:
    """Counts used for EXPECTED_COUNTS cross-checking."""
    path = Path(xlsx_path)
    blocks = 0
    empty = 0
    columns = 0
    unlabeled = 0
    subjects: set[str] = set()
    by_protocol: dict[str, int] = {}
    missing_feature_cells = 0
    seen: set[tuple] = set()
    for b in iter_urfu_blocks(path):
        if b["empty"]:
            empty += 1
            continue
        blocks += 1
        columns += 1
        if b["subject_id"].startswith("UNLABELED_"):
            unlabeled += 1
        subjects.add(b["subject_id"])
        proto = PROTOCOL_MAP[b["sheet"]].value
        by_protocol[proto] = by_protocol.get(proto, 0) + 1
        all_features = {}
        all_features.update(b["features_1"])
        all_features.update(b["features_2"])
        missing_feature_cells += sum(1 for v in all_features.values() if v is None)
        seen.add((b["subject_id"], proto, b["age"]))
    return {
        "blocks": blocks,
        "empty_blocks": empty,
        "signal_columns": columns,
        "unique_subject_columns": len(subjects),
        "unlabeled_columns": unlabeled,
        "by_protocol": by_protocol,
        "missing_feature_cells": missing_feature_cells,
    }
